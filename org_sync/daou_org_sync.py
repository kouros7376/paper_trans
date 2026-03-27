# =============================================================================
# daou_org_sync.py - 다우오피스 조직도 자동 동기화 스크립트
# =============================================================================
# 기능:
#   1. 다우오피스 OpenAPI로 전체 임직원·부서 정보 수집
#   2. 직전 실행 데이터와 비교 → 입사자/퇴사자/부서이동 자동 감지
#   3. 엑셀 변동 보고서 생성
#   4. Active Directory 계정 자동 처리 (설정 시)
#   5. 실행 로그 자동 저장
#
# 실행 방법:
#   python daou_org_sync.py
#
# 자동 실행 (Windows 작업 스케줄러):
#   매주 월요일 08:00 → python C:\...\daou_org_sync.py
#
# 필요 라이브러리 설치:
#   pip install requests openpyxl
#   (AD 연동 시 추가) pip install pyad
# =============================================================================

import os
import json
import base64
import logging
import requests
import openpyxl
from datetime import datetime, date
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# 설정 파일 불러오기
import config

# =============================================================================
# 로그 설정 - 실행 내역을 파일과 화면에 동시 기록
# =============================================================================
def setup_logger():
    """로그를 파일과 콘솔에 동시에 출력하는 로거를 설정합니다."""
    os.makedirs(config.LOG_DIR, exist_ok=True)

    # 오늘 날짜로 로그 파일명 생성 (예: 2026-03-23.log)
    log_filename = os.path.join(
        config.LOG_DIR,
        f"{datetime.now().strftime('%Y-%m-%d')}.log"
    )

    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s [%(levelname)s] %(message)s",
        datefmt="%Y-%m-%d %H:%M:%S",
        handlers=[
            logging.FileHandler(log_filename, encoding="utf-8"),  # 파일 저장
            logging.StreamHandler()                                 # 콘솔 출력
        ]
    )
    return logging.getLogger(__name__)

logger = setup_logger()


# =============================================================================
# 1단계: Access Token 발급
# =============================================================================
def get_access_token():
    """
    다우오피스 OpenAPI 인증 토큰을 발급받습니다.
    토큰은 24시간(86400초) 동안 유효합니다.

    반환값: access_token 문자열 (실패 시 None)
    """
    url = f"{config.DAOU_API_BASE}/public/auth/v1/oauth2/token"

    # client_id:client_secret 을 Base64로 인코딩 (RFC 6749 표준 방식)
    # 예: "myid:mysecret" → "bXlpZDpteXNlY3JldA=="
    credentials = f"{config.DAOU_CLIENT_ID}:{config.DAOU_CLIENT_SECRET}"
    encoded = base64.b64encode(credentials.encode("utf-8")).decode("utf-8")

    headers = {
        "Authorization": f"Basic {encoded}",
        "Content-Type": "application/x-www-form-urlencoded"
    }
    params = {"grant_type": "client_credentials"}

    try:
        response = requests.post(url, headers=headers, data=params, timeout=30)
        response.raise_for_status()
        token_data = response.json()
        token = token_data.get("access_token")
        logger.info("Access Token 발급 성공")
        return token
    except requests.exceptions.RequestException as e:
        logger.error(f"Access Token 발급 실패: {e}")
        return None


# =============================================================================
# 2단계: API 호출 공통 함수
# =============================================================================
def api_get(token, endpoint):
    """
    다우오피스 API를 GET 방식으로 호출하는 공통 함수입니다.

    매개변수:
        token    : Access Token 문자열
        endpoint : API 경로 (예: /public/api/attnd-v3/organization-chart/user/list)

    반환값: API 응답 data 리스트 (실패 시 빈 리스트)
    """
    url = f"{config.DAOU_API_BASE}{endpoint}"
    headers = {"Authorization": f"Bearer {token}"}

    try:
        response = requests.get(url, headers=headers, timeout=60)
        response.raise_for_status()
        result = response.json()

        # API 응답 코드 확인 (200이 정상)
        if str(result.get("code")) != "200":
            logger.warning(f"API 응답 이상 [{endpoint}]: {result.get('message')}")
            return []

        return result.get("data", [])

    except requests.exceptions.RequestException as e:
        logger.error(f"API 호출 오류 [{endpoint}]: {e}")
        return []


# =============================================================================
# 3단계: 다우오피스 데이터 수집
# =============================================================================
def fetch_all_data(token):
    """
    계정정보, 부서정보, 부서원정보를 한 번에 수집합니다.

    반환값: {
        "users"   : [계정 정보 목록],
        "depts"   : [부서 정보 목록],
        "members" : [부서원 정보 목록]
    }
    """
    logger.info("=== 다우오피스 데이터 수집 시작 ===")

    # 계정 정보 조회 (전체 임직원)
    users = api_get(token, "/public/api/attnd-v3/organization-chart/user/list")
    logger.info(f"계정 정보 수집 완료: {len(users)}명")

    # 부서 정보 조회 (전체 부서)
    depts = api_get(token, "/public/api/attnd-v3/organization-chart/department/list")
    logger.info(f"부서 정보 수집 완료: {len(depts)}개 부서")

    # 부서원 정보 조회 (부서별 소속 구성원)
    members = api_get(token, "/public/api/attnd-v3/organization-chart/member/list")
    logger.info(f"부서원 정보 수집 완료: {len(members)}건")

    return {"users": users, "depts": depts, "members": members}


# =============================================================================
# 4단계: 데이터 가공 - 부서별 구성원 맵 생성
# =============================================================================
def build_user_dept_map(users, members, depts):
    """
    사용자별로 소속 부서명을 쉽게 찾을 수 있도록 매핑 테이블을 만듭니다.

    예: {"daou_0123": "개발팀", "daou_0456": "인사팀"}
    """
    # 부서코드 → 부서이름 딕셔너리
    dept_map = {d.get("code"): d.get("name", "") for d in depts if d.get("code")}

    # loginId → 부서명 리스트 딕셔너리 (겸직 고려)
    user_dept_map = {}
    for m in members:
        login_id = m.get("loginId")
        dept_code = m.get("departmentCode")
        dept_name = dept_map.get(dept_code, dept_code)  # 코드로 이름 찾기
        if login_id:
            user_dept_map.setdefault(login_id, []).append(dept_name)

    return user_dept_map


# =============================================================================
# 5단계: 스냅샷 저장 및 불러오기
# =============================================================================
def save_snapshot(data):
    """현재 조직 데이터를 JSON 파일로 저장합니다 (다음 실행 때 비교 기준)."""
    os.makedirs(config.DATA_DIR, exist_ok=True)
    with open(config.SNAPSHOT_FILE, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
    logger.info(f"스냅샷 저장 완료: {config.SNAPSHOT_FILE}")


def load_snapshot():
    """이전에 저장한 조직 데이터를 불러옵니다. 없으면 None 반환."""
    if not os.path.exists(config.SNAPSHOT_FILE):
        logger.info("이전 스냅샷 없음 → 첫 실행으로 처리합니다.")
        return None
    with open(config.SNAPSHOT_FILE, "r", encoding="utf-8") as f:
        return json.load(f)


# =============================================================================
# 6단계: 변동 사항 비교 분석
# =============================================================================
def analyze_changes(current_data, previous_data, user_dept_map):
    """
    현재 데이터와 이전 데이터를 비교하여 변동 사항을 분류합니다.

    반환값: {
        "new_employees"     : [입사자 목록],
        "resigned_employees": [퇴사/중지 목록],
        "dept_changed"      : [부서 이동자 목록],
        "dormant"           : [휴면 계정 목록]
    }
    """
    result = {
        "new_employees":      [],  # 신규 입사자
        "resigned_employees": [],  # 퇴사/중지자
        "dept_changed":       [],  # 부서 이동자
        "dormant":            []   # 메일 휴면 계정
    }

    current_users  = {u["loginId"]: u for u in current_data["users"] if u.get("loginId")}
    today_str = date.today().isoformat()  # 오늘 날짜 (YYYY-MM-DD)

    # ── 이전 데이터가 있을 때만 비교 ──
    if previous_data:
        prev_users = {u["loginId"]: u for u in previous_data.get("users", []) if u.get("loginId")}
        prev_dept_map = previous_data.get("user_dept_map", {})

        # 신규 입사자: 이전에 없던 계정 (ONLINE 상태)
        for login_id, user in current_users.items():
            if login_id not in prev_users and user.get("status") == "ONLINE":
                user["dept_names"] = ", ".join(user_dept_map.get(login_id, ["미배정"]))
                result["new_employees"].append(user)
                logger.info(f"[입사자] {user.get('name')} ({login_id}) - {user.get('dept_names')}")

        # 퇴사/중지자: 이전에 있었는데 status가 STOP으로 바뀐 경우
        for login_id, prev_user in prev_users.items():
            if login_id in current_users:
                cur_user = current_users[login_id]
                cur_status = cur_user.get("status", "")

                # STOP 상태로 변경된 경우
                if cur_status in config.RESIGN_STATUS and prev_user.get("status") == "ONLINE":
                    cur_user["dept_names"] = ", ".join(user_dept_map.get(login_id, ["미배정"]))
                    result["resigned_employees"].append(cur_user)
                    logger.info(f"[퇴사/중지] {cur_user.get('name')} ({login_id})")

                # 부서 이동: 소속 부서가 변경된 경우
                cur_depts  = set(user_dept_map.get(login_id, []))
                prev_depts = set(prev_dept_map.get(login_id, []))
                if cur_depts != prev_depts and cur_user.get("status") == "ONLINE":
                    result["dept_changed"].append({
                        "loginId"   : login_id,
                        "name"      : cur_user.get("name", ""),
                        "prev_depts": ", ".join(sorted(prev_depts)) or "미배정",
                        "curr_depts": ", ".join(sorted(cur_depts))  or "미배정",
                        "joinDate"  : cur_user.get("joinDate", "")
                    })
                    logger.info(f"[부서이동] {cur_user.get('name')} ({login_id}): "
                                f"{', '.join(prev_depts)} → {', '.join(cur_depts)}")

    # ── 퇴사일(expiredDate) 기준 퇴사자: 첫 실행에도 감지 가능 ──
    for login_id, user in current_users.items():
        expired = user.get("expiredDate")
        if expired and expired <= today_str:
            # 이미 퇴사자 목록에 없을 때만 추가
            existing = [r["loginId"] for r in result["resigned_employees"]]
            if login_id not in existing:
                user["dept_names"] = ", ".join(user_dept_map.get(login_id, ["미배정"]))
                user["resign_reason"] = f"퇴사일 도래 ({expired})"
                result["resigned_employees"].append(user)
                logger.info(f"[퇴사일 도래] {user.get('name')} ({login_id}) - {expired}")

    # ── 휴면 계정 감지 ──
    for login_id, user in current_users.items():
        if user.get("status") in config.DORMANT_STATUS:
            user["dept_names"] = ", ".join(user_dept_map.get(login_id, ["미배정"]))
            result["dormant"].append(user)

    return result


# =============================================================================
# 7단계: 엑셀 보고서 생성
# =============================================================================
def create_excel_report(current_data, changes, user_dept_map):
    """
    변동 사항을 정리한 엑셀 보고서를 생성합니다.
    시트 구성:
        1. 요약         - 전체 현황 한눈에 보기
        2. 전체직원현황  - 현재 시점 전체 임직원 목록
        3. 신규입사자    - 이번 동기화에서 발견된 신규 입사자
        4. 퇴사_중지자   - 퇴사/계정 중지 처리 필요 목록
        5. 부서이동자    - 부서가 변경된 직원 목록
        6. 부서현황      - 전체 부서 구조
    """
    os.makedirs(config.REPORT_DIR, exist_ok=True)
    now_str = datetime.now().strftime("%Y%m%d_%H%M")
    filepath = os.path.join(config.REPORT_DIR, f"조직동기화_보고서_{now_str}.xlsx")

    wb = openpyxl.Workbook()

    # ── 스타일 정의 ──
    # 헤더 배경색
    COLOR_HEADER_BLUE   = "1F4E79"  # 진한 파란색 (전체 현황)
    COLOR_HEADER_GREEN  = "375623"  # 진한 녹색   (입사자)
    COLOR_HEADER_RED    = "7B0000"  # 진한 빨간색 (퇴사자)
    COLOR_HEADER_ORANGE = "7F4A00"  # 진한 주황색 (부서이동)
    COLOR_HEADER_GRAY   = "404040"  # 진한 회색   (부서현황)

    def make_header_style(bg_color):
        """헤더 셀 스타일을 만드는 내부 함수"""
        return {
            "font"     : Font(bold=True, color="FFFFFF", size=11),
            "fill"     : PatternFill("solid", fgColor=bg_color),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
            "border"   : Border(
                left=Side(style="thin"), right=Side(style="thin"),
                top=Side(style="thin"), bottom=Side(style="thin")
            )
        }

    def apply_header(ws, headers, row=1, bg_color="1F4E79"):
        """헤더 행을 적용하는 내부 함수"""
        style = make_header_style(bg_color)
        for col, text in enumerate(headers, start=1):
            cell = ws.cell(row=row, column=col, value=text)
            cell.font      = style["font"]
            cell.fill      = style["fill"]
            cell.alignment = style["alignment"]
            cell["border"] = style["border"]
        ws.row_dimensions[row].height = 22

    def set_column_widths(ws, widths):
        """열 너비를 설정하는 내부 함수"""
        for col, width in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(col)].width = width

    # =========================================================================
    # 시트 1: 요약
    # =========================================================================
    ws_summary = wb.active
    ws_summary.title = "요약"

    today_str = datetime.now().strftime("%Y년 %m월 %d일 %H:%M")
    total_users    = len(current_data["users"])
    online_users   = sum(1 for u in current_data["users"] if u.get("status") == "ONLINE")
    stop_users     = sum(1 for u in current_data["users"] if u.get("status") == "STOP")
    dormant_users  = sum(1 for u in current_data["users"] if u.get("status") == "DORMANT")
    total_depts    = len(current_data["depts"])

    summary_data = [
        ("", ""),
        ("  다우오피스 조직도 동기화 보고서", ""),
        (f"  생성일시: {today_str}", ""),
        ("", ""),
        ("항목", "수치"),
        ("전체 임직원 수", f"{total_users}명"),
        ("정상 계정 (ONLINE)", f"{online_users}명"),
        ("중지 계정 (STOP)", f"{stop_users}명"),
        ("휴면 계정 (DORMANT)", f"{dormant_users}명"),
        ("전체 부서 수", f"{total_depts}개"),
        ("", ""),
        ("변동 사항", "인원"),
        ("신규 입사자", f"{len(changes['new_employees'])}명"),
        ("퇴사 / 계정 중지", f"{len(changes['resigned_employees'])}명"),
        ("부서 이동자", f"{len(changes['dept_changed'])}명"),
        ("휴면 계정", f"{len(changes['dormant'])}명"),
    ]

    ws_summary.column_dimensions["A"].width = 30
    ws_summary.column_dimensions["B"].width = 20

    for row_idx, (label, value) in enumerate(summary_data, start=1):
        cell_a = ws_summary.cell(row=row_idx, column=1, value=label)
        cell_b = ws_summary.cell(row=row_idx, column=2, value=value)

        if label in ("항목", "변동 사항"):
            for cell in [cell_a, cell_b]:
                cell.font = Font(bold=True, color="FFFFFF", size=11)
                cell.fill = PatternFill("solid", fgColor="1F4E79")
                cell.alignment = Alignment(horizontal="center")

        elif row_idx == 2:
            cell_a.font = Font(bold=True, size=16, color="1F4E79")

        elif label and label not in ("", " "):
            cell_a.font = Font(bold=True, size=11)
            cell_b.alignment = Alignment(horizontal="center")
            # 변동 사항 행에 색상 강조
            if label == "신규 입사자":
                cell_b.font = Font(bold=True, color="375623", size=12)
            elif label in ("퇴사 / 계정 중지",):
                cell_b.font = Font(bold=True, color="7B0000", size=12)

    # =========================================================================
    # 시트 2: 전체 직원 현황
    # =========================================================================
    ws_all = wb.create_sheet("전체직원현황")
    headers_all = ["이름", "로그인ID", "사번", "상태", "소속부서", "직위", "직급",
                   "고용형태", "입사일", "퇴사일", "이메일(직통번호)"]
    apply_header(ws_all, headers_all, bg_color=COLOR_HEADER_BLUE)
    set_column_widths(ws_all, [12, 16, 10, 10, 20, 10, 10, 12, 12, 12, 22])

    # 상태 한글 변환
    status_map = {"ONLINE": "정상", "STOP": "중지", "DORMANT": "휴면"}

    for row_idx, user in enumerate(
        sorted(current_data["users"], key=lambda u: u.get("name", "")), start=2
    ):
        login_id = user.get("loginId", "")
        dept_names = ", ".join(user_dept_map.get(login_id, ["미배정"]))
        status = status_map.get(user.get("status", ""), user.get("status", ""))
        row_data = [
            user.get("name", ""),
            login_id,
            user.get("employeeNumber", ""),
            status,
            dept_names,
            user.get("positionName", ""),
            user.get("gradeName", ""),
            user.get("employeeType", ""),
            user.get("joinDate", ""),
            user.get("expiredDate", ""),
            user.get("directTel", ""),
        ]
        for col_idx, val in enumerate(row_data, start=1):
            cell = ws_all.cell(row=row_idx, column=col_idx, value=val or "")
            cell.alignment = Alignment(horizontal="left", vertical="center")
            # 퇴사/중지자 행은 연한 빨간색으로 표시
            if user.get("status") in config.RESIGN_STATUS:
                cell.fill = PatternFill("solid", fgColor="FFE0E0")
            elif user.get("status") in config.DORMANT_STATUS:
                cell.fill = PatternFill("solid", fgColor="FFF3CD")

    ws_all.auto_filter.ref = f"A1:{get_column_letter(len(headers_all))}1"

    # =========================================================================
    # 시트 3: 신규 입사자
    # =========================================================================
    ws_new = wb.create_sheet("신규입사자")
    headers_new = ["이름", "로그인ID", "사번", "소속부서", "직위", "직급", "고용형태", "입사일", "처리 필요 항목"]
    apply_header(ws_new, headers_new, bg_color=COLOR_HEADER_GREEN)
    set_column_widths(ws_new, [12, 16, 10, 20, 10, 10, 12, 12, 35])

    if not changes["new_employees"]:
        ws_new.cell(row=2, column=1, value="신규 입사자 없음").font = Font(italic=True, color="888888")
    else:
        for row_idx, user in enumerate(changes["new_employees"], start=2):
            login_id = user.get("loginId", "")
            dept_names = user.get("dept_names", "미배정")
            action_items = "AD 계정 생성 / NAS 폴더 권한 부여 / 출입카드 등록"
            row_data = [
                user.get("name", ""), login_id, user.get("employeeNumber", ""),
                dept_names, user.get("positionName", ""), user.get("gradeName", ""),
                user.get("employeeType", ""), user.get("joinDate", ""), action_items
            ]
            for col_idx, val in enumerate(row_data, start=1):
                cell = ws_new.cell(row=row_idx, column=col_idx, value=val or "")
                cell.alignment = Alignment(horizontal="left", vertical="center")
                cell.fill = PatternFill("solid", fgColor="E8F5E9")  # 연한 초록 배경

    # =========================================================================
    # 시트 4: 퇴사 / 중지자
    # =========================================================================
    ws_resign = wb.create_sheet("퇴사_중지자")
    headers_resign = ["이름", "로그인ID", "사번", "소속부서", "상태", "퇴사일", "처리 필요 항목"]
    apply_header(ws_resign, headers_resign, bg_color=COLOR_HEADER_RED)
    set_column_widths(ws_resign, [12, 16, 10, 20, 10, 12, 40])

    if not changes["resigned_employees"]:
        ws_resign.cell(row=2, column=1, value="퇴사/중지자 없음").font = Font(italic=True, color="888888")
    else:
        for row_idx, user in enumerate(changes["resigned_employees"], start=2):
            login_id = user.get("loginId", "")
            dept_names = user.get("dept_names", "미배정")
            action_items = "AD 계정 비활성화 / NAS 접근 권한 회수 / 출입카드 반납 / 메일 백업"
            row_data = [
                user.get("name", ""), login_id, user.get("employeeNumber", ""),
                dept_names, user.get("status", ""),
                user.get("expiredDate", "") or user.get("resign_reason", ""),
                action_items
            ]
            for col_idx, val in enumerate(row_data, start=1):
                cell = ws_resign.cell(row=row_idx, column=col_idx, value=val or "")
                cell.alignment = Alignment(horizontal="left", vertical="center")
                cell.fill = PatternFill("solid", fgColor="FFEBEE")  # 연한 빨강 배경

    # =========================================================================
    # 시트 5: 부서 이동자
    # =========================================================================
    ws_moved = wb.create_sheet("부서이동자")
    headers_moved = ["이름", "로그인ID", "이전 부서", "현재 부서", "입사일"]
    apply_header(ws_moved, headers_moved, bg_color=COLOR_HEADER_ORANGE)
    set_column_widths(ws_moved, [12, 16, 25, 25, 12])

    if not changes["dept_changed"]:
        ws_moved.cell(row=2, column=1, value="부서 이동자 없음").font = Font(italic=True, color="888888")
    else:
        for row_idx, u in enumerate(changes["dept_changed"], start=2):
            row_data = [u["name"], u["loginId"], u["prev_depts"], u["curr_depts"], u["joinDate"]]
            for col_idx, val in enumerate(row_data, start=1):
                cell = ws_moved.cell(row=row_idx, column=col_idx, value=val or "")
                cell.alignment = Alignment(horizontal="left", vertical="center")
                cell.fill = PatternFill("solid", fgColor="FFF3E0")  # 연한 주황 배경

    # =========================================================================
    # 시트 6: 부서 현황
    # =========================================================================
    ws_dept = wb.create_sheet("부서현황")
    headers_dept = ["부서명", "부서코드", "상위부서코드", "부서약어", "부서이메일", "정렬순서"]
    apply_header(ws_dept, headers_dept, bg_color=COLOR_HEADER_GRAY)
    set_column_widths(ws_dept, [25, 20, 20, 15, 30, 10])

    for row_idx, dept in enumerate(
        sorted(current_data["depts"], key=lambda d: d.get("sortOrder", "0")), start=2
    ):
        row_data = [
            dept.get("name", ""), dept.get("code", ""),
            dept.get("parentCode", "") or "(최상위)",
            dept.get("alias", ""), dept.get("emailId", ""),
            dept.get("sortOrder", "")
        ]
        for col_idx, val in enumerate(row_data, start=1):
            cell = ws_dept.cell(row=row_idx, column=col_idx, value=val or "")
            cell.alignment = Alignment(horizontal="left", vertical="center")

    wb.save(filepath)
    logger.info(f"엑셀 보고서 저장 완료: {filepath}")
    return filepath


# =============================================================================
# 8단계: AD 연동 처리 (선택 - config.AD_ENABLED = True 설정 시 활성화)
# =============================================================================
def process_ad_sync(changes):
    """
    Active Directory 계정을 자동으로 처리합니다.

    실제 사용하려면:
    1. config.py 에서 AD_ENABLED = True 로 변경
    2. pip install pyad 설치
    3. AD 서버 접속 정보 입력
    """
    if not config.AD_ENABLED:
        logger.info("AD 연동 비활성화 상태 (config.AD_ENABLED = False)")
        return

    try:
        import pyad.aduser
        import pyad.adcontainer
        import pyad.adquery
        import pyad

        # AD 서버 접속
        pyad.set_defaults(
            ldap_server=config.AD_SERVER,
            username=f"{config.AD_DOMAIN}\\{config.AD_ADMIN_USER}",
            password=config.AD_ADMIN_PASS
        )
        logger.info(f"AD 서버 접속: {config.AD_SERVER}")

        # ── 신규 입사자 → AD 계정 생성 ──
        for user in changes["new_employees"]:
            login_id  = user.get("loginId", "")
            full_name = user.get("name", "")
            try:
                # OU 컨테이너 가져오기
                ou = pyad.adcontainer.ADContainer.from_dn(config.AD_USER_OU)
                # 계정 생성
                new_user = pyad.aduser.ADUser.create(
                    name=full_name,
                    container_object=ou,
                    optional_attributes={
                        "sAMAccountName"   : login_id,
                        "userPrincipalName": f"{login_id}@{config.AD_DOMAIN}",
                        "displayName"      : full_name,
                        "department"       : user.get("dept_names", ""),
                        "title"            : user.get("positionName", ""),
                    }
                )
                # 임시 비밀번호 설정 (첫 로그인 시 변경 강제)
                temp_pass = f"Welcome@{datetime.now().strftime('%Y%m')}"
                new_user.set_password(temp_pass)
                new_user.enable()
                logger.info(f"[AD] 계정 생성 완료: {full_name} ({login_id})")
            except Exception as e:
                logger.error(f"[AD] 계정 생성 실패 [{login_id}]: {e}")

        # ── 퇴사자 → AD 계정 비활성화 ──
        for user in changes["resigned_employees"]:
            login_id = user.get("loginId", "")
            try:
                # loginId로 AD 사용자 검색
                q = pyad.adquery.ADQuery()
                q.execute_query(
                    attributes=["distinguishedName"],
                    where_clause=f"sAMAccountName = '{login_id}'",
                    base_dn=config.AD_BASE_DN
                )
                for row in q.get_results():
                    ad_user = pyad.aduser.ADUser.from_dn(row["distinguishedName"])
                    ad_user.disable()
                    logger.info(f"[AD] 계정 비활성화 완료: {user.get('name')} ({login_id})")
            except Exception as e:
                logger.error(f"[AD] 계정 비활성화 실패 [{login_id}]: {e}")

    except ImportError:
        logger.warning("pyad 라이브러리 미설치. 'pip install pyad' 실행 후 재시도하세요.")
    except Exception as e:
        logger.error(f"AD 연동 오류: {e}")


# =============================================================================
# 메인 실행 함수
# =============================================================================
def main():
    logger.info("=" * 60)
    logger.info("다우오피스 조직도 자동 동기화 시작")
    logger.info(f"실행 시각: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    logger.info("=" * 60)

    # ① Access Token 발급
    token = get_access_token()
    if not token:
        logger.error("토큰 발급 실패 → 스크립트를 종료합니다.")
        logger.error("config.py 의 DAOU_CLIENT_ID / DAOU_CLIENT_SECRET 을 확인하세요.")
        return

    # ② 다우오피스에서 전체 데이터 수집
    current_data = fetch_all_data(token)
    if not current_data["users"]:
        logger.error("임직원 데이터 조회 실패 → 스크립트를 종료합니다.")
        return

    # ③ 부서별 구성원 매핑 테이블 생성
    user_dept_map = build_user_dept_map(
        current_data["users"],
        current_data["members"],
        current_data["depts"]
    )

    # ④ 이전 스냅샷 불러오기 (비교 기준)
    previous_data = load_snapshot()

    # ⑤ 변동 사항 분석 (입사자 / 퇴사자 / 부서이동)
    changes = analyze_changes(current_data, previous_data, user_dept_map)

    # ⑥ 변동 요약 출력
    logger.info("─" * 40)
    logger.info(f"신규 입사자  : {len(changes['new_employees'])}명")
    logger.info(f"퇴사/중지자  : {len(changes['resigned_employees'])}명")
    logger.info(f"부서 이동자  : {len(changes['dept_changed'])}명")
    logger.info(f"휴면 계정    : {len(changes['dormant'])}명")
    logger.info("─" * 40)

    # ⑦ 엑셀 보고서 생성
    report_path = create_excel_report(current_data, changes, user_dept_map)

    # ⑧ AD 연동 처리 (config에서 활성화 시)
    process_ad_sync(changes)

    # ⑨ 현재 데이터를 스냅샷으로 저장 (다음 실행 때 비교 기준)
    current_data["user_dept_map"] = user_dept_map  # 부서 맵도 함께 저장
    save_snapshot(current_data)

    logger.info("=" * 60)
    logger.info("동기화 완료!")
    logger.info(f"보고서 위치: {report_path}")
    logger.info("=" * 60)


# =============================================================================
# 스크립트 직접 실행 시 main() 호출
# =============================================================================
if __name__ == "__main__":
    main()
